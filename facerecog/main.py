import telepot
import cv2
import time
import datetime as dt
import openpyxl
from openpyxl import Workbook
import os
from PIL import Image
import numpy as np
import glob



command = 'a'
sourceCam = 0


id_message = 1693238711
count_handle = 1
count_handle_prev = 0
isSend = True
isCommand = True
last_people = ""

camss = cv2.VideoCapture(sourceCam)
camss.set(3, 660)
camss.set(4, 500)
camss.release()
ret, frame = camss.read()
# Define min window size to be recognized as a face
minW = 0.1 * camss.get(3)
minH = 0.1 * camss.get(4)

faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
eye = cv2.CascadeClassifier('haarcascade_eye.xml')

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath)

font = cv2.FONT_HERSHEY_SIMPLEX
id = 0
face_id = 0

names = []

first_message = "Welcome to home security\n " \
                "- Manage our security\n "\
                "______________________ \n"\
                "Tell hi ============> /hi,\n"\
                "capture ============> /capture,\n"\
                "Add People   ====> /add_people \n" \
                "Manage People ======> /people \n" \
                "Open the Door   ====> /open \n" \
                "Start smart door ===> /door"
first_send = "Welcome to home security\n " \
                    "- Manage our security\n " \
                    "______________________ \n" \
                    "Tell hi ============> /hi,\n" \
                    "capture ============> /capture,\n" \
                    "Add People   ====> /add_people \n" \
                    "Manage People ======> /people \n" \
                    "Open the Door   ====> /open \n" \
                    "Start smart door ===> /door"

def capture():
    global id_message, first_message, sourceCam
    print("Camera Take")
    cams = cv2.VideoCapture(sourceCam)
    cams.set(3, 660)
    cams.set(4, 500)


    bot.sendMessage(id_message, "camera take, wait a minute...")
    isCapture = True

    while isCapture:
        ret, frame = cams.read()
        t = time.localtime()
        current_time = time.strftime("%H%M%S", t)
        img_name = "Dokumen/capture_" + current_time + ".jpg"
        cv2.imwrite(img_name, frame)
        bot.sendPhoto(id_message, photo=open(img_name, 'rb'))
        isCapture = False
    cams.release()
    bot.sendMessage(id_message, first_send)

def getImagesAndLabels():
    print("aaaa")
    path = 'dataset'
    imagePaths = [os.path.join(path,f) for f in os.listdir(path)]
    faceSamples=[]
    ids = []


    for imagePath in imagePaths:

        PIL_img = Image.open(imagePath).convert('L') # convert it to grayscale
        img_numpy = np.array(PIL_img,'uint8')

        id = int(os.path.split(imagePath)[-1].split(".")[1])
        faces = faceCascade.detectMultiScale(img_numpy)

        for (x,y,w,h) in faces:
            faceSamples.append(img_numpy[y:y+h,x:x+w])
            ids.append(id)

    return faceSamples,ids

def train():
    print("aaa")
    print("\n [INFO] Training faces. It will take a few seconds. Wait ...")
    faces, ids = getImagesAndLabels()
    recognizer.train(faces, np.array(ids))

    # Save the model into trainer/trainer.yml
    recognizer.write('trainer/trainer.yml')  # recognizer.save() worked on Mac, but not on Pi
    print("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))

def tess():
    x = 1
    while x < 101:
        if x % 10 == 0:
            print("x")
        x +=1
        print(x)


def ada_people():
    global id_message, first_message, sourceCam, command
    cam = cv2.VideoCapture(sourceCam)
    cam.set(3, 660)
    cam.set(4, 500)
    isLoop = True
    while isLoop:
        if command == '/break':
            cv2.destroyAllWindows()
            break
        if command == '/close':
            closeDoor()

        ret, img = cam.read()
        img = cv2.flip(img, 1)  # Flip vertically

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale(
            gray,
            scaleFactor=1.2,
            minNeighbors=5,
            minSize=(int(minW), int(minH)),
        )
        if len(faces) >= 1:
            bot.sendMessage(id_message, "good position")
            isLoop = False
        else:
            bot.sendMessage(id_message, "bad position")
            isLoop = False
    cam.release()


def take_people():
    global id_message, first_message, sourceCam, command, face_id
    m = "silahkan berdiri didepan kamera\n " \
        "-               -\n " \
        "/next            /check_camera\n " \
        "-               -\n " \
        "/break "

    capture()
    ada_people()
    bot.sendMessage(id_message, m)
    print(command)

    count = 0
    isName = False
    c = 0
    latest_command =""
    cam = cv2.VideoCapture(sourceCam)
    cam.set(3, 660)
    cam.set(4, 500)
    while (True):
        if latest_command != command:
            print(command)
            print(names)
            latest_command = command
            if command == '/break':
                cv2.destroyAllWindows()
                break
            if command == "/check_camera":
                bot.sendMessage(id_message, "camera take, wait a minute...")
                isCapture = True

                while isCapture:
                    ret, frame = cam.read()
                    t = time.localtime()
                    current_time = time.strftime("%H%M%S", t)
                    img_name = "Dokumen/capture_" + current_time + ".jpg"
                    cv2.imwrite(img_name, frame)
                    bot.sendPhoto(id_message, photo=open(img_name, 'rb'))
                    isCapture = False
                time.sleep(1)

                isLoop = True
                while isLoop:
                    if command == '/break':
                        cv2.destroyAllWindows()
                        break
                    if command == '/close':
                        closeDoor()

                    ret, img = cam.read()
                    img = cv2.flip(img, 1)  # Flip vertically

                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

                    faces = faceCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.2,
                        minNeighbors=5,
                        minSize=(int(minW), int(minH)),
                    )
                    if len(faces) >= 1:
                        bot.sendMessage(id_message, "good position")
                        isLoop = False
                    else:
                        bot.sendMessage(id_message, "bad position")
                        isLoop = False

                bot.sendMessage(id_message, m)
                command = "emt"
            elif command == "/next":
                bot.sendMessage(id_message, "take People Start")

                print("\n [INFO] Camera take people")
                isDone = False

                while not isDone:

                    ret, img = cam.read()
                    img = cv2.flip(img, 1)  # flip video image vertically
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    faces = faceCascade.detectMultiScale(gray, 1.3, 6)

                    for (x, y, w, h) in faces:
                        cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                        count += 1
                        c += 1
                        int(count)
                        int(c)

                        # Save the captured image into the datasets folder
                        cv2.imwrite("dataset/User." + str(face_id) + '.' + str(count) + ".jpg", gray[y:y + h, x:x + w])
                        print(count)
                        cv2.imshow('image', img)
                        if c % 10 == 0:
                            print("x")
                            y = c / 10
                            y = y * 10
                            bot.sendMessage(id_message, " [INFO] Progress on " + str(int(y))+ " %")

                    if count == 100:
                        isDone = True

                    command = "emt"

                # Do a bit of cleanup
                bot.sendMessage(id_message, " [INFO] take picture done")
                bot.sendMessage(id_message, " [INFO] Training Image Start,\n[INFO]  wait a minute...")
                train()
                bot.sendMessage(id_message, "\n [INPUT] name of people :")
                isName = True

            if isName and command != "emt":
                print("command")
                print(command)

                folder = "Datasets/"
                xl = "listPeople.xlsx"
                wb_obj = openpyxl.load_workbook(folder + xl)

                sheet_obj = wb_obj.active
                max_row = sheet_obj.max_row

                book_new = Workbook()
                sheet_new = book_new.active

                isName = False
                names.append(command)
                print("nnnn:")
                l = 1
                for i in names:
                    print(i)
                    sheet_new['A' + str(l)] = str(l)
                    sheet_new['B' + str(l)] = i
                    l += 1
                book_new.save(folder + xl)
    cam.release()





def listPeopple():
    global id_message, command, names, face_id
    folder = "Datasets/"
    xl = "listPeople.xlsx"
    wb_obj = openpyxl.load_workbook(folder + xl)

    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    book_new = Workbook()
    sheet_new = book_new.active
    first_message = "Press if want to delete"
    people = []
    for i in range(1, max_row + 1):

        cell_obj = sheet_obj.cell(row=i, column=2)
        nama = str(cell_obj.value)
        people.append(nama)
        names.append(nama)
        # print(cell_obj.value)
        face_id = i
        first_message = first_message+"\n" \
                        "/" + str(cell_obj.value)
    first_message = first_message.replace(" ", "_")
    print(first_message)
    print(i)
    # bot.sendMessage(id_message, first_message)
    return people
        # x = cell_obj.value.split("?id=")
        # id = x[1]
        # print(id)
        #
        # sheet_new['C' + str(i - 1)] = "aaa"
    # wb_obj.close()
    # book_new.save(folder + xl)



def sendListPeople():
    global id_message, command, names, face_id
    folder = "Datasets/"
    xl = "listPeople.xlsx"
    wb_obj = openpyxl.load_workbook(folder + xl)

    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    book_new = Workbook()
    sheet_new = book_new.active
    first_message = "Press if want to delete"
    people = []
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        nama = str(cell_obj.value)
        people.append(nama)
        names.append(nama)
        # print(cell_obj.value)
        face_id = i
        first_message = first_message + "\n" \
                                        "/" + str(cell_obj.value)
    first_message = first_message.replace(" ", "_")
    first_message = first_message+ "\n\n /break"
    print(first_message)
    print(i)
    bot.sendMessage(id_message, first_message)
    latest_command = ""
    while True:
        if latest_command != command:
            latest_command = command

            commands =  command.split("/")
            commandsd = str(commands[1])
            commandsd = commandsd.replace("_", " ")
            print(commandsd)
            if command == '/break':
                cv2.destroyAllWindows()
                break
            elif commandsd in people:
                print("find ", commandsd)
                isRemove = False

                for i in range(1, max_row + 1):
                    cell_obj = sheet_obj.cell(row=i, column=2)
                    cell_obj_id = sheet_obj.cell(row=i, column=1)
                    nama = str(cell_obj.value)

                    id_people = cell_obj_id.value

                    # cek apakah orang yang dipilih ada
                    if isRemove:

                        print("rename file next")
                        for x in range(1, 101):
                            os.rename("dataset/User."+str(int(id_people) - 1)+"."+str(x)+".jpg", "dataset/User."+str(int(id_people) - 2)+"."+str(x)+".jpg")
                        else:
                            print("Finally finished!")

                    elif commandsd == nama:
                        print("sama : ", commandsd)
                        # hapus dataset orang yang dipilih
                        directory = "dataset"
                        for filename in glob.iglob(f'{directory}/User.'+str(int(id_people) - 1)+'.*'):
                            print(filename)
                            os.remove(filename)
                        isRemove = True
                        people.remove(commandsd)

                    print(nama, " : ", id_people)

                train()

                folder = "Datasets/"
                xl = "listPeople.xlsx"
                os.remove(folder + xl)

                book_new = Workbook()
                sheet_new = book_new.active

                isName = False
                print("nnnn:")
                l = 1
                for i in people:
                    print(i)
                    sheet_new['A' + str(l)] = str(l)
                    sheet_new['B' + str(l)] = i
                    l += 1
                book_new.save(folder + xl)
            else:
                print("not Found")
                print(people)


    first_message = "Welcome to home security\n " \
                    "- Manage our security\n " \
                    "______________________ \n" \
                    "Tell hi ============> /hi,\n" \
                    "capture ============> /capture,\n" \
                    "Add People   ====> /add_people \n" \
                    "Manage People ======> /people \n" \
                    "Open the Door   ====> /open \n" \
                    "Start smart door ===> /door"
    bot.sendMessage(id_message, first_message)

def handle(msg):
    global command, count_handle

    command = msg['text']
    count_handle += 1

    return command

def hi():
    global command, isSend
    bot.sendMessage(id_message, "say hi")
def removePeople():
    print("a")

def openDoor():
    bot.sendMessage(id_message, "Open the Door")
def closeDoor():
    global isSend, isCommand, command
    bot.sendMessage(id_message, "Close the Door")
    isCommand = True
    isSend = True
    command = ""
def doorSystem():
    global isSend, last_people
    bot.sendMessage(id_message, "Smart home run....")
    cam = cv2.VideoCapture(sourceCam)
    cam.set(3, 660)
    cam.set(4, 500)
    while True:
        if command == '/break':
            cv2.destroyAllWindows()
            break
        if command == '/close':
            closeDoor()

        ret, img = cam.read()
        img = cv2.flip(img, 1)  # Flip vertically

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale(
            gray,
            scaleFactor=1.2,
            minNeighbors=6,
            minSize=(int(minW), int(minH)),
        )

        for (x, y, w, h) in faces:

            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

            id, confidence = recognizer.predict(gray[y:y + h, x:x + w])
            if (confidence < 60):
                id_people = names[id]
            else:
                id_people = "unknown"

            print(id_people)
            # Check if confidence is less them 100 ==> "0" is perfect match
            if (confidence < 60):
                if id_people is not last_people:
                    last_people = id_people
                    bot.sendMessage(id_message, "selamat datang " + id_people)
                    openDoor()
                    isSend = True
                # confidence = "  {0}%".format(round(100 - confidence))
            elif id_people is not last_people:
                id_people = "unknown"
                last_people = id_people
                bot.sendMessage(id_message, "Unknown people detected")
                # closeDoor()

            cv2.putText(img, str(id_people), (x + 5, y - 5), font, 1, (255, 255, 255), 2)

        cv2.imshow('camera', img)
        k = cv2.waitKey(1) & 0xff  # Press 'ESC' for exiting video
        if k == 27:
            cam.release()
            break
    cam.release()


aaaaaa = listPeopple()

bot = telepot.Bot('5416163618:AAHBBGGW22gxlbFixf06qlqVCIPqnDUKqtU')
bot.sendMessage(id_message, first_message)
bot.message_loop(handle)

while True:

    if count_handle == 1:
        print('I am Listening....')
        count_handle += 1
        count_handle_prev = 2

    elif count_handle_prev < count_handle:
        print('new syntax')
        print( 'text :', command)
        if command == "/hi":
            a = hi()
        elif command == "/open":
            openDoor()
        elif command == "/door":
            doorSystem()
        elif command == "/people":
            sendListPeople()
        elif command == "/capture":
            capture()
        elif command == "/add_people":
            take_people()
        elif command == "/close":
            closeDoor()

    count_handle_prev = count_handle
